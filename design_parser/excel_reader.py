"""
Excel 文件读取模块。
支持 .xlsx（openpyxl 流式读取）和 .xls（xlrd 读取）。
"""

import openpyxl
import xlrd
from pathlib import Path
from typing import Dict, List

# ==================== 基础读取函数 ====================
def read_excel(file_path: str) -> Dict[str, List[dict]]:
    """读取 Excel 文件所有工作表，返回 {工作表名: [行字典列表]}"""
    path = str(file_path).lower()
    if path.endswith('.xlsx'):
        return _read_xlsx(file_path)
    elif path.endswith('.xls'):
        return _read_xls(file_path)
    else:
        raise ValueError(f"不支持的文件格式: {file_path}")

def _read_xlsx(file_path: str) -> Dict[str, List[dict]]:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    result = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            result[name] = []
            continue
        headers = [str(c).strip() if c is not None else f"Col_{i}" for i, c in enumerate(rows[0])]
        data = []
        for row in rows[1:]:
            data.append({headers[i]: (str(v).strip() if isinstance(v, str) else v) for i, v in enumerate(row)})
        result[name] = data
    wb.close()
    return result

def _read_xls(file_path: str) -> Dict[str, List[dict]]:
    wb = xlrd.open_workbook(file_path)
    result = {}
    for name in wb.sheet_names():
        ws = wb.sheet_by_name(name)
        if ws.nrows == 0:
            result[name] = []
            continue
        headers = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
        data = []
        for r in range(1, ws.nrows):
            row = {}
            for c in range(ws.ncols):
                cell = ws.cell_value(r, c)
                row[headers[c]] = cell if cell != '' else None
            data.append(row)
        result[name] = data
    return result