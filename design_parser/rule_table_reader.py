"""Excel 规则解析器：读取官方《图层表字段说明和数据校验规则.xlsx》。

工作表结构：
- “校验规则” sheet：NO. / 检测项 / 检测内容 / 涉及图层 / 涉及字段 / 规则 / 备注
  （NO. 与子编号为合并单元格，解析时向下填充）
- 其余每个 sheet 为一个图层字段说明：Nom champ / 字段描述 / Descriptif champ /
  Exemple / Type champ / Longueur champ / Format de saisie / Obligatoire O/C/N /
  Description de la condition / Contrainte de remplissage

职责：把官方规则表解析为结构化数据，并转换成规则引擎可消费的
可执行条件描述（check_type + object_type + field + params）。
"""
from pathlib import Path
from typing import List, Dict, Any, Optional
import re

RULES_KEYWORDS = ("规则", "字段说明")
VALIDATION_SHEET_NAMES = ("校验规则", "规则")
EXCEL_EXTS = (".xlsx", ".xls")


def find_rule_files(root: Path) -> List[Path]:
    """递归查找官方规则表 Excel（文件名含“规则”或“字段说明”）。"""
    out = []
    for f in root.rglob("*"):
        if f.is_file() and f.suffix.lower() in EXCEL_EXTS:
            if any(k in f.name for k in RULES_KEYWORDS):
                out.append(f)
    return sorted(out)


def _norm(value: Any) -> str:
    return "" if value is None else str(value).replace("\n", " ").strip()


def _parse_length(raw: str) -> Optional[int]:
    """从官方 Longueur champ 列提取正整数长度（兼容 '30' / '30.0' / 'max 30'）。"""
    m = re.search(r"\d+", raw or "")
    if not m:
        return None
    try:
        return int(m.group(0))
    except (ValueError, TypeError):
        return None


def _sheet_rows(path: Path, sheet: Optional[str] = None):
    """返回指定/第一个工作表的所有行（值列表），xlsx 与 xls 通用。"""
    if path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = next((w for w in wb.worksheets if w.title == sheet), wb.worksheets[0])
            return ws.title, [list(r) for r in ws.iter_rows(values_only=True)]
        finally:
            wb.close()
    import xlrd
    wb = xlrd.open_workbook(str(path))
    sh = next((s for s in wb.sheets() if s.name == sheet), wb.sheets()[0])
    return sh.name, [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]


def read_validation_rules(path: Path, sheet: Optional[str] = None) -> Dict[str, Any]:
    """解析“校验规则”工作表：NO./子编号向下填充，返回结构化规则列表。"""
    sheet_name, rows = _sheet_rows(path, sheet)
    rules = []
    cur_no, cur_sub = "", ""
    for row in rows:
        cells = [_norm(c) for c in row]
        if not any(cells):
            continue
        no = cells[0]
        sub = cells[1]
        item = cells[2]
        content = cells[3]
        layers = cells[4]
        fields = cells[5]
        rule = cells[6]
        remark = cells[7] if len(cells) > 7 else ""
        if no:
            cur_no = no
        if sub:
            cur_sub = sub
        if no == "NO." or item == "检测项":
            continue  # 跳过表头行
        if not (item or content or rule):
            continue
        rules.append({
            "no": cur_no,
            "sub_no": cur_sub,
            "check_item": item,
            "check_content": content,
            "layers": [x.strip() for x in layers.split() if x.strip()],
            "fields": [x.strip() for x in fields.split(",") if x.strip()],
            "rule": rule,
            "remark": remark,
        })
    return {"sheet": sheet_name, "rules": rules}


def read_field_specs(path: Path) -> Dict[str, Any]:
    """解析每个图层 sheet 的字段说明，返回 {图层: [字段字典]}。"""
    result = {}
    for fname in _sheet_filenames(path):
        sheet_name, rows = _sheet_rows(path, fname)
        if sheet_name in VALIDATION_SHEET_NAMES:
            continue
        header_idx = None
        for i, row in enumerate(rows):
            first = _norm(row[0]) if row else ""
            if "nom champ" in first.lower():
                header_idx = i
                break
        if header_idx is None:
            continue
        header = [_norm(c).lower() for c in rows[header_idx]]

        def col(*keys):
            for i, h in enumerate(header):
                if all(k in h for k in keys):
                    return i
            return None

        idx = {
            "name": col("nom champ"),
            "desc_cn": col("中文描述") or col("字段描述"),
            "desc_fr": col("descriptif"),
            "example": col("exemple"),
            "type": col("type", "champ"),
            "length": col("longueur"),
            "format": col("format"),
            "required": col("obligatoire"),
            "condition_desc": col("condition"),
            "constraint": col("contrainte"),
        }
        fields = []
        for row in rows[header_idx + 1:]:
            name = _norm(row[idx["name"]]) if idx["name"] is not None else ""
            if not name:
                continue
            fields.append({
                "name": name,
                "desc_cn": _norm(row[idx["desc_cn"]]) if idx["desc_cn"] is not None else "",
                "desc_fr": _norm(row[idx["desc_fr"]]) if idx["desc_fr"] is not None else "",
                "example": _norm(row[idx["example"]]) if idx["example"] is not None else "",
                "type": _norm(row[idx["type"]]) if idx["type"] is not None else "",
                "length": _norm(row[idx["length"]]) if idx["length"] is not None else "",
                "format": _norm(row[idx["format"]]) if idx["format"] is not None else "",
                "required": _norm(row[idx["required"]]) if idx["required"] is not None else "",
                "condition_desc": _norm(row[idx["condition_desc"]]) if idx["condition_desc"] is not None else "",
                "constraint": _norm(row[idx["constraint"]]) if idx["constraint"] is not None else "",
            })
        result[sheet_name] = fields
    return result


def _sheet_filenames(path: Path):
    if path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()
    import xlrd
    wb = xlrd.open_workbook(str(path))
    return [s.name for s in wb.sheets()]


def build_executable_rules(field_specs: Dict[str, list],
                           validation_rules: Dict[str, Any]) -> List[Dict[str, Any]]:
    """把字段说明与校验规则转换成规则引擎可消费的条件描述。"""
    out: List[Dict[str, Any]] = []

    # 1) 图层字段说明中的必填字段（Obligatoire=O）
    for layer, fields in field_specs.items():
        for f in fields:
            if f["required"].upper() == "O":
                out.append({
                    "check_type": "required_not_empty",
                    "object_type": layer,
                    "field": f["name"],
                    "params": {"required": f["required"]},
                    "source": f"{layer}字段说明",
                })

    # 1.1) 图层字段说明中的类型/长度约束（官方 Type champ / Longueur champ）
    for layer, fields in field_specs.items():
        for f in fields:
            if f["type"]:
                out.append({
                    "check_type": "field_type_check",
                    "object_type": layer,
                    "field": f["name"],
                    "params": {"type": f["type"]},
                    "source": f"{layer}字段说明",
                    "rule_id": "R018",
                })
            length = _parse_length(f["length"])
            if length is not None:
                out.append({
                    "check_type": "field_length_check",
                    "object_type": layer,
                    "field": f["name"],
                    "params": {"length": length, "raw_length": f["length"]},
                    "source": f"{layer}字段说明",
                    "rule_id": "R032",
                })

    # 2) 校验规则表 → 按关键词归类（顺序敏感：先精确后宽泛）
    classifier = [
        ("文件完整性", "file_complete"),
        ("图层完整性", "file_complete"),
        ("图层类型及命名", "layer_type_naming"),
        ("命名规范", "layer_type_naming"),
        ("坐标系一致", "crs_consistent"),
        ("空图层", "layer_not_empty"),
        ("不能重名", "code_unique"),
        ("孤立性", "isolation_bidirectional"),
        ("重叠", "geometry_no_overlap"),
        ("光缆端点", "point_on_cable_endpoint"),
        ("必须位于", "point_in_polygon"),
        ("不得大于", "capacity_limit"),
    ]
    for r in validation_rules.get("rules", []):
        text = r["check_item"] + r["check_content"] + r["rule"]
        check_type = None
        for kw, ct in classifier:
            if kw in text:
                check_type = ct
                break
        if check_type is None:
            check_type = "manual_review"  # 无法自动化的规则标记人工确认
        source = f"{r['no']}.{r['sub_no']} {r['check_item']}".strip(" .")
        out.append({
            "check_type": check_type,
            "object_type": r["layers"][0] if r["layers"] else "",
            "field": r["fields"][0] if r["fields"] else "",
            "params": {
                "layers": r["layers"],
                "fields": r["fields"],
                "rule": r["rule"],
                "content": r["check_content"],
            },
            "source": source,
        })
    return out


def parse_rule_library(path: Path) -> Dict[str, Any]:
    """一键解析官方规则表：校验规则 + 字段说明 + 可执行条件。"""
    validation = read_validation_rules(path)
    field_specs = read_field_specs(path)
    return {
        "file": path.name,
        "validation_rules": validation["rules"],
        "field_specs": field_specs,
        "executable_rules": build_executable_rules(field_specs, validation),
    }
