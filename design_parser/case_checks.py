"""标准测试案例的自动化判定（我们自己定义“标准答案”，对应官方审查规则库）。

案例格式：每个 xlsx 一个 case，工作表 = BOITE/CABLE/IMB/INFRASTRUCTURE/PTECH/
SITE/ZNRO/ZPM/纤芯连接与分配/BOM物料（简化字段）。

判定规则映射：
- 缺少工作表             -> R-FILE-001 文件完整性
- 必填字段 CODE 为空      -> R-FLD-001 必填字段非空
- CODE 重复              -> R-FLD-002 编码唯一
- CABLE 端点引用不存在    -> R-REL-004 光缆端点未连接有效设备
- 设备未被任何光缆引用     -> R-REL-001 孤立设备
- 分路器 OUT 数超过容量   -> R-DAT-001 容量超限（如 2:4 分路器最多 4 个 OUT）
- 无分路器时同一输入纤芯被重复使用 -> R-FIBER-001 纤芯重复占用
- BOM 物料编码不在已知库  -> R-BOM-001 物料无法匹配
"""
import re
from pathlib import Path
from typing import Dict, List, Any

EXPECTED_SHEETS = ["BOITE", "CABLE", "IMB", "INFRASTRUCTURE", "PTECH", "SITE",
                   "ZNRO", "ZPM", "纤芯连接与分配", "BOM物料"]
DEVICE_LAYERS = ["BOITE", "SITE", "PTECH", "IMB", "INFRASTRUCTURE", "ZNRO", "ZPM"]
# CABLE 端点中允许出现的非设备根节点（如 POP 汇聚点）
ROOT_CODE_ALLOWLIST = ("POP", "SRO")


def load_case_sheets(path: Path) -> Dict[str, List[list]]:
    """读取案例 xlsx 的全部工作表为行列表。"""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return {ws.title: [list(r) for r in ws.iter_rows(values_only=True)]
                for ws in wb.worksheets}
    finally:
        wb.close()


def _col(header: list, *names) -> int:
    for i, h in enumerate(header):
        s = str(h).strip()
        if any(s == n or s.startswith(n) for n in names):
            return i
    return -1


def _cell(row: list, idx: int):
    if idx < 0 or idx >= len(row):
        return None
    v = row[idx]
    return None if v is None or str(v).strip() == "" else str(v).strip()


def find_fiber_core_duplicates(sheets: Dict[str, List[list]]) -> List[Dict[str, Any]]:
    """R-FIBER-001：在纤芯连接与分配表中，无分路器时同一输入纤芯被重复使用。

    输入 sheets 为 {工作表名: 行列表}，首行为表头；兼容三种结构：
    - 「所属节点/托盘编号/光分路器/IN」案例表；
    - SRO TOPO 的 SRO Port/ODF Code/ODF Port（同一输入端口重复使用）；
    - 单箱页 Entrée/N°/T/F（同一输入纤芯重复使用）。
    """
    issues: List[Dict[str, Any]] = []
    for rows in sheets.values():
        if not rows:
            continue
        # 官方 SRO TOPO / 单箱页的表头可能不在首行（如 retour/说明行），
        # 在前 8 行内定位真正的表头行，避免官方格式漏检。
        header_idx = 0
        for i in range(min(8, len(rows))):
            r = rows[i]
            if (
                min(_col(r, "所属节点"), _col(r, "托盘编号"), _col(r, "光分路器"), _col(r, "IN")) >= 0
                or min(_col(r, "SRO Port"), _col(r, "ODF Code"), _col(r, "ODF Port")) >= 0
                or min(_col(r, "Entrée"), _col(r, "N°"), _col(r, "T"), _col(r, "F")) >= 0
            ):
                header_idx = i
                break
        rows = rows[header_idx:]
        node_idx = _col(rows[0], "所属节点")
        tray_idx = _col(rows[0], "托盘编号")
        spl_idx = _col(rows[0], "光分路器")
        in_idx = _col(rows[0], "IN")
        if min(node_idx, tray_idx, spl_idx, in_idx) >= 0:
            splice_in = {}
            for i, r in enumerate(rows[1:], 1):
                spl = _cell(r, spl_idx) or ""
                if spl and spl != "无":
                    continue
                node = _cell(r, node_idx) or ""
                tray = _cell(r, tray_idx) or ""
                core = _cell(r, in_idx) or ""
                if not core:
                    continue
                key = (node, tray, core)
                if key in splice_in:
                    issues.append({
                        "rule_id": "R-FIBER-001",
                        "object": f"{node}/{tray}/{core}",
                        "message": f"输入纤芯 {core} 在 {node}/{tray} 被重复使用"
                                   f"（第 {splice_in[key]} 行与第 {i} 行）",
                    })
                else:
                    splice_in[key] = i
            continue
        sro_idx = _col(rows[0], "SRO Port")
        odf_idx = _col(rows[0], "ODF Code")
        odf_port_idx = _col(rows[0], "ODF Port")
        if odf_idx >= 0 and odf_port_idx >= 0:
            used_port = {}
            for i, r in enumerate(rows[1:], 1):
                odf = _cell(r, odf_idx)
                port = _cell(r, odf_port_idx)
                if not odf or not port:
                    continue
                sro = _cell(r, sro_idx) if sro_idx >= 0 else ""
                key = (sro or "", odf, port)
                if key in used_port:
                    issues.append({
                        "rule_id": "R-FIBER-001",
                        "object": f"{odf}/{port}",
                        "message": f"输入端口 {odf}/{port}（SRO {sro or '-'}）被重复使用"
                                   f"（第 {used_port[key]} 行与第 {i} 行）",
                    })
                else:
                    used_port[key] = i
            continue
        ent_idx = _col(rows[0], "Entrée")
        n_idx = _col(rows[0], "N°")
        t_idx = _col(rows[0], "T")
        f_idx = _col(rows[0], "F")
        if min(ent_idx, n_idx, t_idx, f_idx) >= 0:
            used_core = {}
            for i, r in enumerate(rows[1:], 1):
                ent = _cell(r, ent_idx)
                if not ent:
                    continue
                key = (ent, _cell(r, n_idx) or "", _cell(r, t_idx) or "", _cell(r, f_idx) or "")
                if key in used_core:
                    issues.append({
                        "rule_id": "R-FIBER-001",
                        "object": f"{ent}/{key[1]}-{key[2]}-{key[3]}",
                        "message": f"输入纤芯 {ent} N°{key[1]} T{key[2]} F{key[3]} 被重复使用"
                                   f"（第 {used_core[key]} 行与第 {i} 行）",
                    })
                else:
                    used_core[key] = i
            continue
    return issues


def check_case(path: Path, known_material_codes: set = None) -> List[Dict[str, Any]]:
    """对单个案例执行全部判定，返回 issue 列表。"""
    sheets = load_case_sheets(path)
    issues: List[Dict[str, Any]] = []

    # 1) 文件完整性：缺少工作表
    for s in EXPECTED_SHEETS:
        if s not in sheets:
            issues.append({"rule_id": "R-FILE-001", "object": s,
                           "message": f"缺少工作表 {s}"})

    def rows_of(name):
        return sheets.get(name, [])

    # 设备 code 集合
    device_codes = set()
    for layer in DEVICE_LAYERS:
        rows = rows_of(layer)
        if not rows:
            continue
        idx = _col(rows[0], "CODE")
        for r in rows[1:]:
            c = _cell(r, idx)
            if c:
                device_codes.add(c)

    # 2) 必填字段 CODE 为空（R-FLD-001）与 3) CODE 重复（R-FLD-002）
    for layer in ["BOITE", "CABLE", "IMB", "INFRASTRUCTURE", "PTECH", "SITE", "ZNRO", "ZPM"]:
        rows = rows_of(layer)
        if not rows:
            continue
        idx = _col(rows[0], "CODE")
        seen = {}
        for i, r in enumerate(rows[1:], 1):
            c = _cell(r, idx)
            if c is None:
                issues.append({"rule_id": "R-FLD-001", "object": f"{layer} R{i}",
                               "message": f"{layer} 第 {i + 1} 行 CODE 为空"})
            else:
                if c in seen:
                    issues.append({"rule_id": "R-FLD-002", "object": f"{layer} CODE={c}",
                                   "message": f"{layer} CODE {c} 重复（第 {seen[c]} 行与第 {i + 1} 行）"})
                else:
                    seen[c] = i + 1

    # 4) CABLE 端点引用不存在（R-REL-004）
    cable_rows = rows_of("CABLE")
    if cable_rows:
        code_idx = _col(cable_rows[0], "CODE")
        up_idx = _col(cable_rows[0], "起点", "ORIGINE")
        down_idx = _col(cable_rows[0], "终点", "EXTREMITE")
        for r in cable_rows[1:]:
            code = _cell(r, code_idx)
            for side, idx in (("起点", up_idx), ("终点", down_idx)):
                ref = _cell(r, idx)
                if ref is None:
                    continue
                if ref in device_codes:
                    continue
                if ref.upper().startswith(ROOT_CODE_ALLOWLIST):
                    continue
                issues.append({"rule_id": "R-REL-004", "object": f"CABLE {code}",
                               "message": f"光缆 {code} {side} 引用的设备 {ref} 不存在"})

    # 5) 孤立设备：BOITE 未被任何光缆引用（R-REL-001）
    if cable_rows:
        refs = set()
        up_idx = _col(cable_rows[0], "起点", "ORIGINE")
        down_idx = _col(cable_rows[0], "终点", "EXTREMITE")
        for r in cable_rows[1:]:
            for idx in (up_idx, down_idx):
                c = _cell(r, idx)
                if c:
                    refs.add(c)
        boite_rows = rows_of("BOITE")
        if boite_rows:
            b_idx = _col(boite_rows[0], "CODE")
            for r in boite_rows[1:]:
                c = _cell(r, b_idx)
                if c and c not in refs:
                    issues.append({"rule_id": "R-REL-001", "object": f"BOITE {c}",
                                   "message": f"光箱 {c} 未被任何光缆引用（孤立设备）"})

    # 6) 容量超限：分路器 OUT 数 > 容量（R-DAT-001）
    fiber_rows = rows_of("纤芯连接与分配")
    if fiber_rows:
        node_idx = _col(fiber_rows[0], "所属节点")
        tray_idx = _col(fiber_rows[0], "托盘编号")
        spl_idx = _col(fiber_rows[0], "光分路器")
        in_idx = _col(fiber_rows[0], "IN")
        out_idx = _col(fiber_rows[0], "OUT")
        groups = {}
        for i, r in enumerate(fiber_rows[1:], 1):
            node = _cell(r, node_idx) or ""
            tray = _cell(r, tray_idx) or ""
            spl = _cell(r, spl_idx) or ""
            out = _cell(r, out_idx) or ""
            key = (node, tray, spl)
            groups.setdefault(key, []).append((out, i))
        for (node, tray, spl), outs in groups.items():
            nums = re.findall(r"\d+", spl)
            capacity = int(nums[-1]) if nums else 0
            distinct = {o for o, _ in outs}
            if capacity and len(distinct) > capacity:
                issues.append({"rule_id": "R-DAT-001", "object": f"{node}/{tray}/{spl}",
                               "message": f"分路器 {spl} 输出 {len(distinct)} 路，超过容量 {capacity}"})

        # 7) 纤芯重复占用：无分路器时同一输入被重复使用（R-FIBER-001）
        issues.extend(find_fiber_core_duplicates(sheets))

    # 8) BOM 物料无法匹配（R-BOM-001）
    bom_rows = rows_of("BOM物料")
    if bom_rows and known_material_codes:
        mc_idx = _col(bom_rows[0], "Material Code", "物料编码")
        for r in bom_rows[1:]:
            c = _cell(r, mc_idx)
            if c and c not in known_material_codes:
                issues.append({"rule_id": "R-BOM-001", "object": f"BOM {c}",
                               "message": f"物料编码 {c} 不在已知物料库中"})

    return issues


def baseline_material_codes(baseline_path: Path) -> set:
    """从“正确工程案例”的 BOM物料 表提取已知物料编码集合。"""
    sheets = load_case_sheets(baseline_path)
    rows = sheets.get("BOM物料", [])
    if not rows:
        return set()
    idx = _col(rows[0], "Material Code", "物料编码")
    return {_cell(r, idx) for r in rows[1:] if _cell(r, idx)}
