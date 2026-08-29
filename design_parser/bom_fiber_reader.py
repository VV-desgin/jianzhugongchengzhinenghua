"""BOM / 纤芯数据解析：统一读取 Excel 表格与 GPKG 矢量图层为结构化数据。

用于 BOM 生成与纤芯分配工作流。真实数据形态：
- BOM_LIST.xlsx / material_code_*.xls（BOM 物料清单与物料编码库）
- SRO-*-TOPO_*.xlsx（纤芯拓扑表）
- BOX / CABLE / SRO.gpkg（纤芯矢量图层）

约定：
- 表格统一输出 {"file", "kind", "sheets": [{"name", "row_count", "rows"}]}
- 行数据统一输出 {"file", "sheet", "headers", "rows"}，首行为表头；
- 超大表（如 BOM_LIST 104 万行）只读取前 limit 行，避免内存与响应膨胀；
- 过滤/翻页按流式单遍扫描，内存只保留当前页行，不随全表增长。
"""
import os
from pathlib import Path
from typing import List, Dict, Any, Optional

BOM_KEYWORDS = ("bom", "material", "list", "物料", "编码")
FIBER_KEYWORDS = ("fiber", "fibre", "topo", "splice", "纤芯", "接续", "分配",
                  "上游", "下游", "upstream", "downstream")
EXCEL_EXTS = (".xlsx", ".xls")
_FILE_CACHE: Dict[str, tuple] = {}
_CACHE_MAX = 256


def read_sheet_rows_multi(path: Path, sheets: List[str],
                          limit: int = 100) -> Dict[str, Any]:
    """一次打开 Excel 文件读取多个 sheet（避免重复打开文件，SRO TOPO 多页签场景）。

    返回 {sheet_name: {"headers": [...], "rows": [[...]], "total": n}}；
    与 read_sheet_rows 单表结构的 headers/rows 兼容（省略 file/kind/page 字段）。
    结果按（文件+sheet列表+limit+修改时间）缓存。
    """
    wanted = [s for s in sheets if s]
    if not wanted:
        return {}
    key = f"multi:{sorted(wanted)}:{limit}"

    def build() -> Dict[str, Any]:
        out: Dict[str, Any] = {}
        suffix = path.suffix.lower()
        if suffix == ".xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                wanted_set = set(wanted)
                for ws in wb.worksheets:
                    if ws.title not in wanted_set:
                        continue
                    raw: List[list] = []
                    for row in ws.iter_rows(values_only=True):
                        raw.append([_to_serializable(c) for c in row])
                        if len(raw) >= HEADER_SCAN_ROWS + limit:
                            break
                    header_idx = _detect_header_index(raw)
                    out[ws.title] = {
                        "headers": raw[header_idx] if raw else [],
                        "rows": raw[header_idx + 1:][:limit],
                        "total": ws.max_row or 0,
                    }
            finally:
                wb.close()
        else:
            import xlrd
            wb = xlrd.open_workbook(str(path))
            wanted_set = set(wanted)
            for sh in wb.sheets():
                if sh.name not in wanted_set:
                    continue
                raw: List[list] = []
                for r_idx in range(min(sh.nrows, HEADER_SCAN_ROWS + limit)):
                    raw.append([_to_serializable(sh.cell_value(r_idx, c)) for c in range(sh.ncols)])
                header_idx = _detect_header_index(raw)
                out[sh.name] = {
                    "headers": raw[header_idx] if raw else [],
                    "rows": raw[header_idx + 1:][:limit],
                    "total": sh.nrows,
                }
        return out

    return _cached_file(path, key, build)


def _cached_file(path: Path, key: str, builder):
    """按（路径+键+修改时间）缓存解析结果，文件变动后自动重算；容量超限时淘汰最旧条目。"""
    try:
        mtime = path.stat().st_mtime
    except OSError:
        mtime = 0
    ck = (str(path), key, mtime)
    hit = _FILE_CACHE.get(ck)
    if hit is not None:
        return hit
    result = builder()
    _FILE_CACHE[ck] = result
    if len(_FILE_CACHE) > _CACHE_MAX:
        for stale in list(_FILE_CACHE)[: _CACHE_MAX // 2]:
            _FILE_CACHE.pop(stale, None)
    return result


DEFAULT_ROW_LIMIT = 100
HEADER_SCAN_ROWS = 10


def _detect_header_index(raw_rows: List[list]) -> int:
    """自动识别真实表头行：多级合并表头（如 SRO TOPO）前几行为元信息。
    优先取“无数字单元格最多”的行（表头多为纯文本列名），兼容合并表头占位空格。
    """
    if not raw_rows:
        return 0
    scan = raw_rows[:HEADER_SCAN_ROWS]
    total_counts = [
        sum(1 for c in r if c is not None and str(c).strip() != "")
        for r in scan
    ]
    first = total_counts[0]
    mx_total = max(total_counts)
    if first >= 5 or first == mx_total:
        return 0
    alpha_counts = [
        sum(1 for c in r if c is not None and str(c).strip() and not any(ch.isdigit() for ch in str(c).strip()))
        for r in scan
    ]
    mx_alpha = max(alpha_counts)
    if mx_alpha >= 5 and mx_alpha >= first * 2:
        return alpha_counts.index(mx_alpha)
    if mx_total - first >= 5 and mx_total >= first * 2:
        return total_counts.index(mx_total)
    return 0


def classify_table(name: str) -> str:
    """按文件名判断表格类型：bom / fiber / other。"""
    low = name.lower()
    if any(k in low for k in FIBER_KEYWORDS):
        return "fiber"
    if any(k in low for k in BOM_KEYWORDS):
        return "bom"
    return "other"


def find_excel_files(root: Path, kinds: tuple = ("bom", "fiber")) -> List[Path]:
    """递归查找指定类型的 Excel 文件（.xlsx/.xls）。"""
    out = []
    for f in root.rglob("*"):
        if f.is_file() and f.suffix.lower() in EXCEL_EXTS:
            if classify_table(f.name) in kinds:
                out.append(f)
    return sorted(out)


def find_gpkg_files(root: Path) -> List[Path]:
    """递归查找 GPKG 矢量文件（纤芯 BOX/CABLE/SRO 等）。"""
    return sorted(f for f in root.rglob("*.gpkg") if f.is_file())


def _to_serializable(value: Any) -> Any:
    """把 Excel/矢量属性值转为可 JSON 序列化的基础类型。"""
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    if hasattr(value, "isoformat"):  # datetime/date
        return value.isoformat()
    return str(value)


def _xlsx_sheets(path: Path) -> List[Dict[str, Any]]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return [{"name": ws.title, "row_count": ws.max_row or 0} for ws in wb.worksheets]
    finally:
        wb.close()


def _xls_sheets(path: Path) -> List[Dict[str, Any]]:
    import xlrd
    wb = xlrd.open_workbook(str(path))
    return [{"name": sh.name, "row_count": sh.nrows} for sh in wb.sheets()]


def list_sheet_names(path: Path) -> List[str]:
    """返回 Excel 文件所有工作表名（只读元数据，适合超大文件）。"""
    if path.suffix.lower() == ".xlsx":
        return [s["name"] for s in _xlsx_sheets(path)]
    return [s["name"] for s in _xls_sheets(path)]


def workbook_summary(path: Path, row_limit: int = DEFAULT_ROW_LIMIT) -> Dict[str, Any]:
    """返回 Excel 工作簿摘要（缓存：按文件+修改时间；单次打开工作簿遍历全部页签）。"""
    def build():
        suffix = path.suffix.lower()
        sheets: List[Dict[str, Any]] = []
        if suffix == ".xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                for ws in wb.worksheets:
                    raw_rows: List[list] = []
                    for row in ws.iter_rows(values_only=True):
                        raw_rows.append([_to_serializable(c) for c in row])
                        if len(raw_rows) >= HEADER_SCAN_ROWS + row_limit:
                            break
                    header_idx = _detect_header_index(raw_rows)
                    sheets.append({
                        "name": ws.title,
                        "row_count": ws.max_row or 0,
                        "rows": raw_rows[header_idx + 1:][:row_limit],
                        "headers": raw_rows[header_idx] if raw_rows else [],
                    })
            finally:
                wb.close()
        else:
            import xlrd
            wb = xlrd.open_workbook(str(path))
            for sh in wb.sheets():
                raw_rows = []
                for r_idx in range(sh.nrows):
                    raw_rows.append([_to_serializable(sh.cell_value(r_idx, c)) for c in range(sh.ncols)])
                    if len(raw_rows) >= HEADER_SCAN_ROWS + row_limit:
                        break
                header_idx = _detect_header_index(raw_rows)
                sheets.append({
                    "name": sh.name,
                    "row_count": sh.nrows,
                    "rows": raw_rows[header_idx + 1:][:row_limit],
                    "headers": raw_rows[header_idx] if raw_rows else [],
                })
        return {"file": path.name, "kind": classify_table(path.name), "sheets": sheets}
    return _cached_file(path, f"summary:{row_limit}", build)


def read_sheet_rows(path: Path, sheet: Optional[str] = None,
                    limit: int = DEFAULT_ROW_LIMIT, filter: Optional[str] = None,
                    page: int = 1, page_size: Optional[int] = None) -> Dict[str, Any]:
    """读取指定工作表数据，自动识别真实表头（兼容多级合并表头）。

    支持：
    - limit/（page+page_size）分页；默认页大小为 limit（上限 1000）；
    - filter：字段值包含关键字（忽略大小写）。
    大数据量优化：流式单遍扫描，内存只保留当前页行；filter/翻页时才全表扫描，
    但不再把整张表构建为列表（BOM_LIST 104 万行场景内存安全）。
    结果按（文件+参数+修改时间）缓存，重复请求不重扫。
    """
    page = max(1, int(page))
    page_size = max(1, min(int(page_size if page_size is not None else limit), 1000))
    suffix = path.suffix.lower()
    kw = filter.strip().lower() if filter else None
    start_idx = (page - 1) * page_size
    key = f"rows:{sheet}:{limit}:{filter}:{page}:{page_size}"

    def build() -> Dict[str, Any]:
        header_rows: List[list] = []
        header_idx = 0
        headers: list = []
        scan_done = False
        total = 0
        page_rows: List[list] = []
        sheet_name = sheet or ""

        def feed(row: list) -> bool:
            """处理一行：返回 True 表示可以提前结束。"""
            nonlocal header_idx, scan_done, total
            if not scan_done:
                header_rows.append(row)
                if len(header_rows) >= HEADER_SCAN_ROWS:
                    header_idx = _detect_header_index(header_rows)
                    headers[:] = header_rows[header_idx]
                    scan_done = True
                    for extra in header_rows[header_idx + 1:]:
                        if feed(extra):
                            return True
                return False
            if kw is not None:
                if not any(kw in str(v).lower() for v in row if v is not None):
                    return False
            total += 1
            if total > start_idx and len(page_rows) < page_size:
                page_rows.append(row)
            if kw is None and page == 1 and len(page_rows) >= page_size:
                return True
            return False

        if suffix == ".xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                ws = next((w for w in wb.worksheets if w.title == sheet), None)
                if ws is None:
                    return {"file": path.name, "kind": classify_table(path.name), "sheet": sheet,
                            "headers": [], "rows": [], "total": 0, "page": page,
                            "page_size": page_size, "sheet_missing": True}
                sheet_name = ws.title
                for row in ws.iter_rows(values_only=True):
                    if feed([_to_serializable(c) for c in row]):
                        break
            finally:
                wb.close()
        else:
            import xlrd
            wb = xlrd.open_workbook(str(path))
            sh = next((s for s in wb.sheets() if s.name == sheet), None)
            if sh is None:
                return {"file": path.name, "kind": classify_table(path.name), "sheet": sheet,
                        "headers": [], "rows": [], "total": 0, "page": page,
                        "page_size": page_size, "sheet_missing": True}
            sheet_name = sh.name
            for r_idx in range(sh.nrows):
                cells = [_to_serializable(sh.cell_value(r_idx, c)) for c in range(sh.ncols)]
                if feed(cells):
                    break
        if not scan_done and header_rows:
            header_idx = _detect_header_index(header_rows)
            headers[:] = header_rows[header_idx]
            scan_done = True
            for extra in header_rows[header_idx + 1:]:
                feed(extra)
        return {"file": path.name, "kind": classify_table(path.name),
                "sheet": sheet_name, "headers": headers, "rows": page_rows,
                "total": total, "page": page, "page_size": page_size}

    return _cached_file(path, key, build)


def gpkg_summary(path: Path) -> Dict[str, Any]:
    """返回 GPKG 矢量图层摘要：图层名、要素数、字段清单、前 DEFAULT_ROW_LIMIT 行。"""
    import fiona
    with fiona.open(path) as src:
        fields = list(src.schema["properties"].keys())
        count = len(src)
        rows = []
        for i, feat in enumerate(src):
            if i >= DEFAULT_ROW_LIMIT:
                break
            rows.append([_to_serializable(feat["properties"].get(f)) for f in fields])
        return {
            "file": path.name,
            "layer": src.name,
            "count": count,
            "fields": fields,
            "rows": rows,
        }


def read_gpkg_rows(path: Path, limit: int = DEFAULT_ROW_LIMIT) -> Dict[str, Any]:
    """读取 GPKG 图层前 limit 行（与 read_sheet_rows 结构一致）。"""
    import fiona
    limit = max(1, min(int(limit), 1000))
    with fiona.open(path) as src:
        fields = list(src.schema["properties"].keys())
        rows = []
        for i, feat in enumerate(src):
            if i >= limit:
                break
            rows.append([_to_serializable(feat["properties"].get(f)) for f in fields])
        return {"file": path.name, "kind": "fiber", "sheet": src.name,
                "headers": fields, "rows": rows}

def collect_code_column(path: Path, sheet: Optional[str] = None,
                        keywords: tuple = ("CODE", "编码"),
                        max_codes: int = 0) -> set:
    """流式全量扫描 Excel 编码列，返回去重编码集合（不截断行数，内存安全）。

    - sheet 为 None 时读取第一个工作表（与 read_sheet_rows 语义一致）；
    - 表头自动识别（兼容多级合并表头），编码列取第一个命中 keywords 的列；
    - 空值/空白忽略；max_codes>0 时最多收集 max_codes 个去重码，防止异常文件撑爆内存。
    """
    max_codes = max(0, int(max_codes))
    key = f"codes:{sheet}:{keywords}:{max_codes}"

    def build() -> set:
        codes: set = set()
        header_rows: List[list] = []
        header_idx = 0
        scan_done = False
        col = 0

        def find_col(headers: list) -> int:
            for i, h in enumerate(headers):
                s = str(h).upper() if h is not None else ""
                if any(k.upper() in s for k in keywords):
                    return i
            return -1  # 未命中编码关键词时不猜测列，避免把名称列当编码收集

        def feed(row: list) -> bool:
            """处理一行，返回 True 表示已收集够可以提前结束。"""
            nonlocal header_idx, scan_done, col
            if not scan_done:
                header_rows.append(row)
                if len(header_rows) >= HEADER_SCAN_ROWS:
                    header_idx = _detect_header_index(header_rows)
                    col = find_col(header_rows[header_idx])
                    scan_done = True
                    for extra in header_rows[header_idx + 1:]:
                        if feed(extra):
                            return True
                return False
            if col < 0:
                return False
            v = row[col] if col < len(row) else None
            if v is not None and str(v).strip():
                codes.add(str(v).strip())
            return bool(max_codes) and len(codes) >= max_codes

        suffix = path.suffix.lower()
        if suffix == ".xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                ws = wb.worksheets[0] if not sheet else next((w for w in wb.worksheets if w.title == sheet), None)
                if ws is None:
                    return set()
                for row in ws.iter_rows(values_only=True):
                    if feed([_to_serializable(c) for c in row]):
                        break
            finally:
                wb.close()
        else:
            import xlrd
            wb = xlrd.open_workbook(str(path))
            sh = wb.sheets()[0] if not sheet else next((s for s in wb.sheets() if s.name == sheet), None)
            if sh is None:
                return set()
            for r_idx in range(sh.nrows):
                cells = [_to_serializable(sh.cell_value(r_idx, c)) for c in range(sh.ncols)]
                if feed(cells):
                    break
        if not scan_done and header_rows:
            header_idx = _detect_header_index(header_rows)
            col = find_col(header_rows[header_idx])
            scan_done = True  # 表尾补扫：小表不足 HEADER_SCAN_ROWS 行时，剩余行应进入收集而非继续累积表头
            for extra in header_rows[header_idx + 1:]:
                if feed(extra):
                    break
        return codes

    return _cached_file(path, key, build)
