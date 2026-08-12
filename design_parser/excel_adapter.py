"""纯 Excel 工程包适配器。

目标：让含官方图层 Sheet 的 xlsx/xls 直接进入主总控，无需 QGIS/SHP。
- 识别 8 个官方图层 Sheet（BOITE/CABLE/PTECH/SITE/INFRASTRUCTURE/ZNRO/ZPM/IMB），大小写/前后缀容错；
- 表头字段别名统一（ENGINEERING_OBJECTS 源字段别名 + 标准中文表头）；
- 行转 UnifiedFeature（geometry=None），供 engineering_data 与非空间规则使用；
- 无空间坐标：由调用方跳过 GIS 空间规则（见 api.py RULE_ROUTING["Excel 工程包"]）。
"""
import re
from pathlib import Path
from typing import Dict, List, Optional

from .feature import UnifiedFeature
from .project_data import ENGINEERING_OBJECTS

OFFICIAL_SHEETS = ("BOITE", "CABLE", "PTECH", "SITE", "INFRASTRUCTURE", "ZNRO", "ZPM", "IMB")

# 标准中文/常见表头 → 官方字段（与 case_checks 口径一致）
HEADER_ALIASES = {
    "编码": "CODE", "编号": "CODE", "代码": "CODE",
    "起点": "ORIGINE", "上游": "ORIGINE",
    "终点": "EXTREMITE", "下游": "EXTREMITE",
    "容量": "CAPACITE", "类型": "TYPE", "型号": "TYPE",
    "数量": "NB_FIBRE_UTIL", "已用纤芯": "NB_FIBRE_UTIL", "纤芯数": "NB_FIBRE_UTIL",
    "长度": "LONGUEUR", "高度": "HAUTEUR_APPUI",
    "MATERIAL CODE": "CODE", "物料编码": "CODE", "物料名称": "NAME", "数量": "QTY",
    "归属PM": "REF_PM", "所属PM": "REF_PM", "所属NRO": "REF_NRO",
    "状态": "STATUT", "生产商": "FABRIQUANT", "产品编号": "REF_PRODUIT",
}


def _alias_lookup() -> Dict[str, str]:
    """构建 表头（大写）→ 官方字段 的别名映射。"""
    lookup = {}
    for _obj, fields in ENGINEERING_OBJECTS.items():
        for _out, src_keys in fields.items():
            for k in src_keys:
                lookup[k.upper()] = k.upper()
    for zh, official in HEADER_ALIASES.items():
        lookup[zh.upper()] = official
    return lookup


_ALIAS = _alias_lookup()


def _normalize_header(value: str, layer: Optional[str] = None) -> str:
    text = str(value).strip().upper()
    if not text:
        return ""
    if layer == "CABLE" and text == "类型":
        return "TYPE_CABLE"
    if text in _ALIAS:
        return _ALIAS[text]
    return text


def _match_sheet(name: str) -> Optional[str]:
    """Sheet 名 → 官方图层名（含后缀/大小写容错），参考表（l_/Type）返回 None。"""
    n = name.strip().upper()
    if n.startswith(("L_", "TYPE")):
        return None
    for official in OFFICIAL_SHEETS:
        if n == official or n.startswith(official + "_") or n.endswith("_" + official) or official in n:
            return official
    return None


def excel_has_official_sheets(path: Path) -> bool:
    """判断 Excel 文件是否包含至少一个官方图层 Sheet。"""
    try:
        from .bom_fiber_reader import list_sheet_names
        for name in list_sheet_names(path):
            if _match_sheet(name):
                return True
    except Exception:
        return False
    return False

def excel_has_bom_sheet(path: Path) -> bool:
    """判断 Excel 文件是否包含 BOM物料 附表 Sheet（可独立识别为 Excel 工程包）。"""
    try:
        from .bom_fiber_reader import list_sheet_names
        for name in list_sheet_names(path):
            if "物料" in name.strip().upper():
                return True
    except Exception:
        return False
    return False


def _iter_sheet_rows(path: Path, sheet_name: str):
    """按 Sheet 名产出（表头行, 数据行迭代器）。"""
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = next((w for w in wb.worksheets if w.title == sheet_name), None)
            if ws is None:
                return None, []
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            return rows
        finally:
            wb.close()
    else:
        import xlrd
        wb = xlrd.open_workbook(str(path))
        sh = next((s for s in wb.sheets() if s.name == sheet_name), None)
        if sh is None:
            return None, []
        rows = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
        return rows


def _value(v):
    if v is None:
        return None
    if isinstance(v, (str, int, float, bool)):
        return v
    if hasattr(v, "isoformat"):
        return v.isoformat()
    return str(v)


_EXTRA_LAYER_SHEETS = {"BOM物料": "BOM物料"}


def _build_layers_from_file(path: Path) -> Dict[str, List[UnifiedFeature]]:
    from .bom_fiber_reader import list_sheet_names
    layers: Dict[str, List[UnifiedFeature]] = {}
    for raw_name in list_sheet_names(path):
        official = _match_sheet(raw_name)
        if official is None and raw_name.strip().upper() in {k.upper() for k in _EXTRA_LAYER_SHEETS}:
            official = raw_name.strip()
        if official is None:
            continue
        rows = _iter_sheet_rows(path, raw_name)
        if not rows:
            continue
        header_row = None
        header_idx = -1
        for i, row in enumerate(rows[:5]):
            if any(c is not None and str(c).strip() for c in row):
                header_row = [_normalize_header(c, official) for c in row]
                header_idx = i
                break
        if header_row is None:
            continue
        features = []
        for i, row in enumerate(rows[header_idx + 1:], 1):
            if all(c is None or str(c).strip() == "" for c in row):
                continue
            props = {}
            for col, h in enumerate(header_row):
                if not h:
                    continue
                if col < len(row):
                    v = _value(row[col])
                    props[h] = "" if v is None or str(v).strip() == "" else v
            features.append(UnifiedFeature(
                source_layer_name=raw_name,
                feature_id=f"EXCEL-{official}-{i}",
                geometry=None,
                properties=props,
                original_crs=None,
            ))
        if features:
            layers.setdefault(official, []).extend(features)
    return layers


def find_excel_project_files(root: Path) -> List[Path]:
    return sorted(p for p in root.rglob("*")
                  if p.is_file() and p.suffix.lower() in (".xlsx", ".xls"))


def build_excel_layers(root: Path) -> Dict[str, List[UnifiedFeature]]:
    """从目录（解压根/单文件所在目录）中构建 Excel 图层；无官方 Sheet 时返回空。"""
    layers: Dict[str, List[UnifiedFeature]] = {}
    for f in find_excel_project_files(root):
        try:
            part = _build_layers_from_file(f)
        except Exception:
            continue
        for official, feats in part.items():
            layers.setdefault(official, []).extend(feats)
    return layers
