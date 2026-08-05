"""规程知识库解析器：读取官方《施工规程知识库.xlsx》为结构化检索数据。

工作表列（与数据方交付格式一致）：
施工对象 / 工序名称 / 操作步骤 / 使用材料 / 工艺要求 / 测试要求 /
安全要求 / 验收标准 / 常见错误 / 页码及章节来源

供施工指令 Agent / 集成方按施工对象或工序关键词检索。
"""
from pathlib import Path
from typing import List, Dict, Any, Optional

PROCEDURE_KEYWORDS = ("规程", "知识库")
EXCEL_EXTS = (".xlsx", ".xls")
COLUMNS = ["施工对象", "工序名称", "操作步骤", "使用材料", "工艺要求",
           "测试要求", "安全要求", "验收标准", "常见错误", "页码及章节来源"]


def find_procedure_files(root: Path) -> List[Path]:
    """递归查找施工规程知识库 Excel（文件名含“规程”或“知识库”）。"""
    out = []
    for f in root.rglob("*"):
        if f.is_file() and f.suffix.lower() in EXCEL_EXTS:
            if any(k in f.name for k in PROCEDURE_KEYWORDS):
                out.append(f)
    return sorted(out)


def _norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def read_procedure_kb(path: Path) -> Dict[str, Any]:
    """解析规程知识库工作表，返回结构化条目列表。"""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()
    if not rows:
        return {"file": path.name, "entries": []}

    header = [_norm(c) for c in rows[0]]
    idx = {}
    for col in COLUMNS:
        try:
            idx[col] = header.index(col)
        except ValueError:
            idx[col] = -1

    entries = []
    for row in rows[1:]:
        entry = {}
        for col in COLUMNS:
            i = idx[col]
            if i >= 0 and i < len(row) and _norm(row[i]):
                entry[col] = _norm(row[i])
        if entry:
            entries.append(entry)
    return {"file": path.name, "entries": entries}


def search_procedure_kb(path: Path, keyword: str = "") -> Dict[str, Any]:
    """按关键词检索规程条目（匹配任一列内容）。"""
    data = read_procedure_kb(path)
    kw = keyword.strip().lower()
    if not kw:
        return data
    hit = []
    for e in data["entries"]:
        if any(kw in str(v).lower() for v in e.values()):
            hit.append(e)
    return {"file": data["file"], "entries": hit}
