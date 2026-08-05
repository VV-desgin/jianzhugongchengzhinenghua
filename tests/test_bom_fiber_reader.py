"""BOM / 纤芯表格解析接口测试。"""

import io
import tempfile
import zipfile
from pathlib import Path

from openpyxl import Workbook

from design_parser.bom_fiber_reader import (
    classify_table,
    find_excel_files,
    read_sheet_rows,
    workbook_summary,
)


def _make_xlsx(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"
    ws.append(["物料编码", "物料名称", "单位", "数量"])
    ws.append(["500003800", "Pengamanan & Persiapan", "PC", 1])
    ws.append(["500003890", "Transportasi", "PC", 1])
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["x"])
    wb.save(path)
    wb.close()


def test_classify_table():
    assert classify_table("BOM_LIST.xlsx") == "bom"
    assert classify_table("material_code_2026.xls") == "bom"
    assert classify_table("SRO-TOPO_20251212.xlsx") == "fiber"
    assert classify_table("纤芯分配表.xlsx") == "fiber"
    assert classify_table("普通表.xlsx") == "other"


def test_workbook_summary_and_read(tmp_path):
    xlsx = tmp_path / "BOM_LIST.xlsx"
    _make_xlsx(xlsx)
    summary = workbook_summary(xlsx, row_limit=10)
    assert summary["file"] == "BOM_LIST.xlsx"
    assert summary["kind"] == "bom"
    names = [s["name"] for s in summary["sheets"]]
    assert "BOM" in names and "Sheet2" in names
    bom = next(s for s in summary["sheets"] if s["name"] == "BOM")
    assert bom["row_count"] == 3
    assert bom["rows"][0][0] == "500003800"
    assert bom["headers"] == ["物料编码", "物料名称", "单位", "数量"]

    data = read_sheet_rows(xlsx, sheet="BOM", limit=5)
    assert data["headers"] == ["物料编码", "物料名称", "单位", "数量"]
    assert data["rows"][0][1] == "Pengamanan & Persiapan"


def _upload_tables_pkg(client):
    tmp = tempfile.mkdtemp(prefix="tables_pkg_")
    xlsx = Path(tmp) / "BOM_LIST.xlsx"
    _make_xlsx(xlsx)
    fiber = Path(tmp) / "fiber_plan.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "fibre"
    ws.append(["CABLE_CODE", "CAPACITE", "NB_FIBRE_U"])
    ws.append(["CDI-001", 24, 10])
    wb.save(fiber)
    wb.close()

    zip_path = Path(tmp) / "tables_pkg.zip"
    with zipfile.ZipFile(zip_path, "w") as z:
        z.write(xlsx, "BOM_LIST.xlsx")
        z.write(fiber, "fiber_plan.xlsx")

    with zip_path.open("rb") as f:
        resp = client.post("/project/load", files={"file": ("tables_pkg.zip", f, "application/zip")})
    assert resp.status_code == 200, resp.text
    return resp.json()["data"]["project_id"]


def test_bom_tables_endpoint(client):
    pid = _upload_tables_pkg(client)
    r = client.get(f"/project/{pid}/bom-tables")
    assert r.status_code == 200
    body = r.json()
    assert body["success"] is True
    files = body["data"]["files"]
    assert len(files) == 1 and files[0]["file"] == "BOM_LIST.xlsx"
    assert any(s["name"] == "BOM" for s in files[0]["sheets"])


def test_fiber_tables_endpoint(client):
    pid = _upload_tables_pkg(client)
    r = client.get(f"/project/{pid}/fiber-tables")
    assert r.status_code == 200
    body = r.json()
    wbs = body["data"]["workbooks"]
    assert len(wbs) == 1 and wbs[0]["file"] == "fiber_plan.xlsx"
    assert wbs[0]["sheets"][0]["name"] == "fibre"


def test_table_data_endpoint(client):
    pid = _upload_tables_pkg(client)
    r = client.get(f"/project/{pid}/table-data", params={"file": "BOM_LIST.xlsx", "sheet": "BOM"})
    assert r.status_code == 200
    data = r.json()["data"]
    assert data["sheet"] == "BOM"
    assert data["headers"] == ["物料编码", "物料名称", "单位", "数量"]
    assert data["rows"][0][0] == "500003800"

    r2 = client.get(f"/project/{pid}/table-data", params={"file": "不存在的表.xlsx"})
    assert r2.status_code == 404


def test_pipeline_include_tables_structure(client, survey_zip_path):
    """data-pipeline include_tables=true 时返回 bom_tables/fiber_tables 固定结构。"""
    with survey_zip_path.open("rb") as f:
        resp = client.post(
            "/agent/data-pipeline",
            files={"file": ("场勘设计图.zip", f, "application/zip")},
            data={"include_tables": "true"},
        )
    assert resp.status_code == 200
    data = resp.json()
    assert data["bom_tables"] == {"files": []}
    assert data["fiber_tables"] == {"workbooks": [], "vectors": []}


def test_read_sheet_rows_filter_and_page(tmp_path):
    xlsx = tmp_path / "BOM_LIST.xlsx"
    _make_xlsx(xlsx)
    # 追加几行便于筛选/分页
    from openpyxl import load_workbook
    wb = load_workbook(xlsx)
    ws = wb["BOM"]
    ws.append(["MAT-002", "Transportasi", 2])
    ws.append(["MAT-003", "Kabel", 3])
    wb.save(xlsx); wb.close()

    filtered = read_sheet_rows(xlsx, sheet="BOM", filter="PC", page=1, page_size=1)
    assert filtered["total"] == 2  # 两行含 PC
    assert len(filtered["rows"]) == 1
    assert filtered["rows"][0][0] == "500003800"

    page2 = read_sheet_rows(xlsx, sheet="BOM", filter="PC", page=2, page_size=1)
    assert len(page2["rows"]) == 1
    assert page2["rows"][0][0] == "500003890"


def test_workbook_summary_cache(tmp_path):
    """解析缓存：同文件同参数命中同一对象，修改时间变化后重算。"""
    import os
    import time

    xlsx = tmp_path / "BOM_LIST.xlsx"
    _make_xlsx(xlsx)
    s1 = workbook_summary(xlsx, row_limit=10)
    s2 = workbook_summary(xlsx, row_limit=10)
    assert s1 is s2  # 命中缓存
    future = time.time() + 10
    os.utime(xlsx, (future, future))
    s3 = workbook_summary(xlsx, row_limit=10)
    assert s3 is not s1  # mtime 变化 -> 重算


def _make_sro_topo_xlsx(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "SRO-JAD-MAR-0001"
    ws.append(["retour", "SRO-JAD-MAR-0001"])
    ws.append(["SRO", None, None, "Type Epissure", None, "Distribution ", None, None])
    ws.append(["SRO Port", "ODF Code", "ODF Port", None, None, "Section", "Code", "Capacité", "N°", "T", "F"])
    ws.append(["1", "ODF01", "1", "E", None, "0001", "CDI-JAD-MAR-0001", "144FO", "1", "1", "1"])
    ws.append(["1", "ODF01", "2", "E", None, "0002", "CDI-JAD-MAR-0001", "144FO", "2", "1", "1"])
    wb.save(path)
    wb.close()


def test_sro_topo_multi_header_detection(tmp_path):
    xlsx = tmp_path / "SRO-TOPO_20251212.xlsx"
    _make_sro_topo_xlsx(xlsx)
    data = read_sheet_rows(xlsx, sheet="SRO-JAD-MAR-0001", limit=10)
    assert data["headers"][0] == "SRO Port"
    assert data["headers"][6] == "Code"
    assert data["rows"][0][0] == "1"
    assert data["rows"][1][1] == "ODF01"
    summary = workbook_summary(xlsx, row_limit=5)
    sh = summary["sheets"][0]
    assert sh["headers"][0] == "SRO Port"
    assert sh["rows"][0][0] == "1"
