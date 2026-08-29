"""P1-1：R-BOM-001 深埋异常码全量扫描测试（评测集 TC-11，BOM_LIST 104 万行场景）。

旧实现三处 limit=5000（_collect_material_codes / _load_official_material_codes /
_find_bom_excel_codes），异常码埋在 5000 行之后会漏检；修复后流式全量扫表。
"""
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook

from design_parser import rule_engine
from design_parser.bom_fiber_reader import collect_code_column, read_sheet_rows


def _make_bom_list(path: Path, rows: int = 5005, bad_code: str = "500099999",
                   bad_row: int = 5006) -> Path:
    """构造 5000+ 行 BOM_LIST：默认前 5005 行合法码，第 5006 行埋异常码。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["物料编码", "物料描述", "单位", "数量"])
    for i in range(1, rows + 1):
        ws.append([f"5000000{i:03d}", "物料", "PC", 1])
    ws.append([bad_code, "深埋异常物料", "PC", 1])
    wb.save(path)
    wb.close()
    return path


def _fake_ctx(root: Path):
    return SimpleNamespace(
        layers={},
        package=SimpleNamespace(temp_dir=str(root)),
    )


def test_collect_code_column_finds_code_after_5000_rows(tmp_path):
    """流式全量扫描能取到 5000 行之后的编码，旧 limit=5000 取不到。"""
    xlsx = _make_bom_list(tmp_path / "BOM_LIST.xlsx")
    codes = collect_code_column(xlsx, sheet="Sheet1")
    assert "500099999" in codes
    old = read_sheet_rows(xlsx, sheet="Sheet1", limit=5000)
    assert "500099999" not in [str(r[0]) for r in old["rows"]]
    # 旧接口 page_size 上限 1000：limit=5000 实际也只返回前 1000 行
    assert len(old["rows"]) == 1000


def test_collect_code_column_small_bom_table(tmp_path):
    """回归：不足 HEADER_SCAN_ROWS(10) 行的小 BOM 表也必须收集编码（表尾补扫未置 scan_done 导致漏检）。"""
    xlsx = tmp_path / "BOM物料.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM物料"
    ws.append(["物料编码", "物料描述", "单位", "数量"])
    ws.append(["500099999", "异常物料", "PC", 1])
    wb.save(xlsx)
    wb.close()
    codes = collect_code_column(xlsx, sheet="BOM物料")
    assert codes == {"500099999"}


def test_collect_code_column_no_code_keyword_returns_empty(tmp_path):
    """表头无任何编码关键词时，不得把第 0 列（如名称）当编码收集。"""
    xlsx = tmp_path / "无编码表.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["名称", "数量"])
    ws.append(["ADSS光缆", 2])
    ws.append(["钢绞线", 5])
    wb.save(xlsx)
    wb.close()
    codes = collect_code_column(xlsx, sheet="Sheet1")
    assert codes == set()


def test_r_bom_001_detects_deep_code(tmp_path):
    """check_bom_material_match 检出埋在 5000 行之后的异常码，且 severity=error。"""
    root = tmp_path / "pkg"
    root.mkdir()
    _make_bom_list(root / "BOM_LIST.xlsx")
    issues = rule_engine.check_bom_material_match(_fake_ctx(root))
    hits = [i for i in issues if i.rule_id == "R-BOM-001" and i.actual_value == "500099999"]
    assert len(hits) == 1
    assert hits[0].passed is False
    assert hits[0].severity == "error"


def test_r_bom_001_dedupes_repeated_bad_code(tmp_path):
    """同一异常码重复 5000+ 次只报 1 条，不会因全量扫描产生海量重复问题。"""
    root = tmp_path / "pkg2"
    root.mkdir()
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["物料编码", "物料描述", "单位", "数量"])
    for _ in range(5005):
        ws.append(["500099999", "重复异常物料", "PC", 1])
    wb.save(root / "BOM_LIST.xlsx")
    wb.close()
    issues = rule_engine.check_bom_material_match(_fake_ctx(root))
    hits = [i for i in issues if i.rule_id == "R-BOM-001"]
    assert len(hits) == 1


def test_load_official_material_codes_full_scan(tmp_path, monkeypatch):
    """官方物料库加载同样全量扫描（5000 行之后的官方码不丢）。"""
    lib = _make_bom_list(tmp_path / "official.xlsx", rows=5005, bad_code="500004729", bad_row=5006)
    monkeypatch.setattr(rule_engine, "OFFICIAL_BOM_LIST_PATH", lib)
    codes = rule_engine._load_official_material_codes()
    assert "500004729" in codes


class _Feat:
    def __init__(self, props):
        self.properties = props


def test_r_bom_001_flags_nonstandard_box_type(tmp_path):
    """非标箱体类型（TC-11：HUAWEI-UNKNOWN-9999-X）→ R-BOM-001 error，即使 BOM 表无异常码。"""
    root = tmp_path / "pkg3"
    root.mkdir()
    ctx = SimpleNamespace(
        layers={"BOITE": [_Feat({"CODE": "PBO-01", "TYPE": "HUAWEI-UNKNOWN-9999-X"})]},
        package=SimpleNamespace(temp_dir=str(root)),
    )
    issues = rule_engine.check_bom_material_match(ctx)
    hits = [i for i in issues if i.rule_id == "R-BOM-001"]
    assert len(hits) == 1
    assert hits[0].problem_location == "TYPE"
    assert hits[0].severity == "error"
    assert "HUAWEI-UNKNOWN-9999-X" in hits[0].error_description


def test_r_bom_001_accepts_standard_box_type(tmp_path):
    """标准箱体类型（PBO/FDT/16口）不触发 R-BOM-001 类型拦截。"""
    root = tmp_path / "pkg4"
    root.mkdir()
    ctx = SimpleNamespace(
        layers={"BOITE": [
            _Feat({"CODE": "PBO-01", "TYPE": "PBO"}),
            _Feat({"CODE": "FDT-01", "TYPE": "FDT"}),
            _Feat({"CODE": "B-16", "TYPE": "16口"}),
        ]},
        package=SimpleNamespace(temp_dir=str(root)),
    )
    issues = rule_engine.check_bom_material_match(ctx)
    assert not [i for i in issues if i.rule_id == "R-BOM-001"]
