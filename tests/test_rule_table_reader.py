"""Excel 规则解析器测试：官方《图层表字段说明和数据校验规则.xlsx》结构解析。"""

import tempfile
import zipfile
from pathlib import Path

from openpyxl import Workbook

from design_parser.rule_table_reader import (
    build_executable_rules,
    find_rule_files,
    parse_rule_library,
    read_field_specs,
    read_validation_rules,
)


def _make_rules_xlsx(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "校验规则"
    ws.append(["NO.", "NO.", "检测项", "检测内容", "涉及图层", "涉及字段", "规则", "备注"])
    ws.append(["1", "1.1", "文件完整性检查", "图层完整性校验", "", "", "shape目录下是否包含全部8个图层文件"])
    ws.append(["", "1.2", "图层类型及命名规范性校验", "", "BOITE", "", "类型：点图层，命名规范：以BOITE结尾"])
    ws.append(["2", "2", "坐标系一致性检查", "QGIS文件和待检查图层的坐标系必须一致"])
    ws.append(["3", "3", "空图层检查", "不能出现空数据的图层"])
    ws.append(["4", "4.1", "图层字段检查", "字段存在，且值不能为空", "BOITE", "CODE,REF_PLAQUE"])
    ws.append(["", "4.2", "", "同一个图层内，该字段不能重名", "BOITE", "CODE"])

    ws2 = wb.create_sheet("BOITE")
    ws2.append(["OBJET : BOITIER", "", "", "", "", "", "", "", "", ""])
    ws2.append(["Nom champ", "字段描述", "Descriptif champ", "Exemple", "Type\nchamp",
                "Longueur\nchamp", "Format de saisie", "Obligatoire\nO/C/N",
                "Description de la condition", "Contrainte de remplissage"])
    ws2.append(["CODE", "光箱编码", "Code", "BPE-001", "Texte", "30", "Saisie libre", "O", "", ""])
    ws2.append(["NOM", "现场命名", "Nom", "BPE-001", "Texte", "30", "Saisie libre", "C", "", ""])
    ws2.append(["CAPACITE", "容量", "Capacite", "144", "Entier", "3", "Saisie libre", "O", "", ""])
    wb.save(path)
    wb.close()


def test_find_rule_files(tmp_path):
    _make_rules_xlsx(tmp_path / "图层表字段说明和数据校验规则.xlsx")
    (tmp_path / "BOM_LIST.xlsx").touch()
    files = find_rule_files(tmp_path)
    assert len(files) == 1
    assert "规则" in files[0].name


def test_read_validation_rules(tmp_path):
    p = tmp_path / "图层表字段说明和数据校验规则.xlsx"
    _make_rules_xlsx(p)
    data = read_validation_rules(p)
    rules = data["rules"]
    assert len(rules) == 6  # 表头行被跳过
    assert rules[0]["no"] == "1" and rules[0]["sub_no"] == "1.1"
    # 合并单元格向下填充
    assert rules[1]["no"] == "1" and rules[1]["sub_no"] == "1.2"
    assert rules[4]["layers"] == ["BOITE"]
    assert rules[5]["no"] == "4" and rules[5]["sub_no"] == "4.2"


def test_read_field_specs_and_executable(tmp_path):
    p = tmp_path / "图层表字段说明和数据校验规则.xlsx"
    _make_rules_xlsx(p)
    specs = read_field_specs(p)
    assert set(specs.keys()) == {"BOITE"}
    assert len(specs["BOITE"]) == 3
    boite = {f["name"]: f for f in specs["BOITE"]}
    assert boite["CODE"]["required"] == "O"
    assert boite["CODE"]["type"] == "Texte"
    assert boite["CAPACITE"]["required"] == "O"

    validation = read_validation_rules(p)
    exec_rules = build_executable_rules(specs, validation)
    types = {r["check_type"] for r in exec_rules}
    assert "required_not_empty" in types  # CODE/CAPACITE 必填
    assert "code_unique" in types        # 4.2 不能重名
    assert "file_complete" in types
    assert "layer_type_naming" in types
    assert "crs_consistent" in types
    assert "layer_not_empty" in types


def test_rule_library_endpoint(client):
    tmp = tempfile.mkdtemp(prefix="rules_pkg_")
    xlsx = Path(tmp) / "图层表字段说明和数据校验规则.xlsx"
    _make_rules_xlsx(xlsx)
    zip_path = Path(tmp) / "rules_pkg.zip"
    with zipfile.ZipFile(zip_path, "w") as z:
        z.write(xlsx, xlsx.name)
    with zip_path.open("rb") as f:
        resp = client.post("/project/load", files={"file": ("rules_pkg.zip", f, "application/zip")})
    assert resp.status_code == 200, resp.text
    pid = resp.json()["data"]["project_id"]

    r = client.get(f"/project/{pid}/rule-library")
    assert r.status_code == 200
    body = r.json()
    assert body["success"] is True
    data = body["data"]
    assert data["file"].endswith(".xlsx")
    assert len(data["validation_rules"]) == 6
    assert set(data["field_specs"].keys()) == {"BOITE"}
    assert any(e["check_type"] == "required_not_empty" for e in data["executable_rules"])

    # 未包含规则表的项目返回空结构而非报错
    r2 = client.get("/project/00000000/rule-library")
    assert r2.status_code == 404  # 项目不存在


def test_parse_rule_library_official_structure(tmp_path):
    """用与官方一致的结构做端到端解析（含 IMB 变体表头）。"""
    p = tmp_path / "图层表字段说明和数据校验规则.xlsx"
    _make_rules_xlsx(p)
    wb = Workbook()
    ws = wb.active
    ws.title = "IMB"
    ws.append(["OBJET : IMMEUBLE"])
    ws.append(["Nom\nchamp", "中文描述", "Descriptif\nchamp", "Exemple", "Type\nchamp",
               "Longueur\nchamp", "Format/valeur attendu(e)", "Obligatoire\nO/C/N",
               "Description de la condition", "Contrainte de remplissage"])
    ws.append(["CODE", "地址唯一编码", "Code", "CAS-001", "Texte", "30", "Saisie libre", "O", "", ""])
    wb.save(p)
    wb.close()
    lib = parse_rule_library(p)
    assert "IMB" in lib["field_specs"]
    assert lib["field_specs"]["IMB"][0]["name"] == "CODE"
    assert lib["field_specs"]["IMB"][0]["desc_cn"] == "地址唯一编码"
    assert lib["field_specs"]["IMB"][0]["required"] == "O"
